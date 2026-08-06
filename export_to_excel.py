# export_to_excel.py
"""
Genera un reporte de Excel (reporte_leads_jofra.xlsx) con el estado de cada lead
dividido en 4 hojas según su etapa en el Kanban.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import BASE_DIR
from database.crud import LeadStatus  # noqa: F401
from database.db import session_scope
from database.models import Lead

logger = logging.getLogger(__name__)

OUTPUT_PATH = str(BASE_DIR / "reporte_leads.xlsx")

# Columnas del reporte
def _get_status_str(status) -> str:
    """Extrae el valor string de un status (sea Enum con .value o directamente string)."""
    if hasattr(status, "value"):
        return str(status.value)
    return str(status) if status is not None else ""


COLUMNS: list[tuple[str, int, callable]] = [
    ("ID", 6, lambda l: l.id),
    ("Empresa", 32, lambda l: l.company_name or ""),
    ("Contacto", 24, lambda l: l.contact_name or ""),
    ("Puesto", 22, lambda l: l.contact_role or ""),
    ("Sector", 20, lambda l: l.sector or ""),
    ("Ciudad", 16, lambda l: l.city or ""),
    ("Correo", 30, lambda l: l.email or ""),
    ("Contexto Icebreaker", 40, lambda l: l.icebreaker_context or ""),
    ("Probabilidad", 16, lambda l: l.purchase_probability or ""),
    ("Estado", 18, lambda l: _get_status_str(l.status)),
    ("Correo enviado", 14, lambda l: _SI_NO(l.email_sent)),
    ("Contestó", 12, lambda l: _SI_NO(l.has_replied)),
    ("Reunión agendada", 16, lambda l: _SI_NO(l.meeting_scheduled)),
    ("Actualizado", 18, lambda l: l.updated_at.strftime("%Y-%m-%d %H:%M") if getattr(l, "updated_at", None) else ""),
]

_HEADER_FILL = PatternFill("solid", fgColor="0B2545")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_SI_FILL = PatternFill("solid", fgColor="D6F5E3")
_NO_FILL = PatternFill("solid", fgColor="F3F4F6")
_MEET_FILL = PatternFill("solid", fgColor="FFE9A8")
_VERY_HIGH_PROB = PatternFill("solid", fgColor="A5D6A7") # Muy verde
_HIGH_PROB = PatternFill("solid", fgColor="C8E6C9")      # Tantito verde
_MED_PROB = PatternFill("solid", fgColor="FFF59D")       # Amarillo
_LOW_PROB = PatternFill("solid", fgColor="FFCC80")       # Tantito rojo (Naranja)
_VERY_LOW_PROB = PatternFill("solid", fgColor="EF9A9A")  # Muy rojo


def _setup_sheet(ws, title):
    ws.title = title
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def export_leads_to_excel(output_path: str = OUTPUT_PATH, ids: list[int] = None) -> str:
    """Construye el archivo Excel con 4 pestañas y devuelve la ruta donde se guardó."""
    wb = Workbook()
    ws_ready = wb.active
    _setup_sheet(ws_ready, "Listos para Enviar")
    ws_conv = wb.create_sheet("En Conversación")
    _setup_sheet(ws_conv, "En Conversación")
    ws_meet = wb.create_sheet("Reunión Agendada")
    _setup_sheet(ws_meet, "Reunión Agendada")
    ws_discard = wb.create_sheet("Rechazados")
    _setup_sheet(ws_discard, "Rechazados")

    with session_scope() as db:
        query = db.query(Lead)
        if ids:
            query = query.filter(Lead.id.in_(ids))
        leads = query.order_by(Lead.created_at.desc()).all()
        
        rows = {
            "Listos para Enviar": 2,
            "En Conversación": 2,
            "Reunión Agendada": 2,
            "Rechazados": 2
        }

        for lead in leads:
            status_val = _get_status_str(lead.status)
            if status_val in ("ready_for_outreach", "new", "sent", "qualifying"):
                sheet_name = "Listos para Enviar"
                ws = ws_ready
            elif status_val in ("in_conversation", "replied"):
                sheet_name = "En Conversación"
                ws = ws_conv
            elif status_val in ("meeting_scheduled", "converted"):
                sheet_name = "Reunión Agendada"
                ws = ws_meet
            elif status_val in ("discarded", "rejected"):
                sheet_name = "Rechazados"
                ws = ws_discard
            else:
                sheet_name = "Listos para Enviar"
                ws = ws_ready

            r = rows[sheet_name]
            for col_idx, (_, _, extractor) in enumerate(COLUMNS, start=1):
                value = extractor(lead)
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")
                header = COLUMNS[col_idx - 1][0]
                
                if header in ("Correo enviado", "Contestó"):
                    cell.fill = _SI_FILL if value == "Sí" else _NO_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header == "Reunión agendada":
                    cell.fill = _MEET_FILL if value == "Sí" else _NO_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header == "Probabilidad":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    val = value.lower()
                    if val == "muy alta":
                        cell.fill = _VERY_HIGH_PROB
                    elif val == "alta" or val == "high":
                        cell.fill = _HIGH_PROB
                    elif val == "media" or val == "medium":
                        cell.fill = _MED_PROB
                    elif val == "baja" or val == "low":
                        cell.fill = _LOW_PROB
                    elif val == "muy baja":
                        cell.fill = _VERY_LOW_PROB
                    else:
                        cell.fill = _MED_PROB
                elif header == "Estado":
                    if "agendada" in value.lower() or "conversaci" in value.lower():
                        cell.fill = _HIGH_PROB
                    elif "descartado" in value.lower():
                        cell.fill = _LOW_PROB
                    else:
                        cell.fill = _MED_PROB
            rows[sheet_name] += 1

    wb.save(output_path)
    logger.info("Reporte de Excel generado: %s", output_path)
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    path = export_leads_to_excel()
    print(f"Reporte generado en: {path}")
