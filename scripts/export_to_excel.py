# export_to_excel.py
"""
Genera un reporte de Excel (reporte_leads.xlsx) con el estado de cada lead, para
que el cliente (Jofra) vea de un vistazo el embudo de Cold Email:

  Quiénes son · Correo · ¿Se les envió correo? · ¿Contestaron? · ¿Agendaron?

Uso:
    python export_to_excel.py
    # o desde el server:  POST /admin/export-excel
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import BASE_DIR
from database.crud import LeadStatus  # noqa: F401  (asegura import del modulo)
from database.db import session_scope
from database.models import Lead

logger = logging.getLogger(__name__)

OUTPUT_PATH = str(BASE_DIR / "reporte_leads.xlsx")

def _get_status_str(status) -> str:
    """Extrae el valor string de un status (sea Enum con .value o directamente string)."""
    if hasattr(status, "value"):
        return str(status.value)
    return str(status) if status is not None else ""


# Columnas del reporte: (encabezado, ancho, extractor(lead) -> valor)
_SI_NO = lambda v: "Sí" if v else "No"  # noqa: E731

COLUMNS: list[tuple[str, int, callable]] = [
    ("ID", 6, lambda l: l.id),
    ("Empresa", 32, lambda l: l.company_name or ""),
    ("Contacto", 24, lambda l: l.contact_name or ""),
    ("Puesto", 22, lambda l: l.contact_role or ""),
    ("Sector", 20, lambda l: l.sector or ""),
    ("Ciudad", 16, lambda l: l.city or ""),
    ("Correo", 30, lambda l: l.email or ""),
    ("Estado", 18, lambda l: _get_status_str(l.status)),
    ("Correo enviado", 14, lambda l: _SI_NO(l.email_sent)),
    ("Contestó", 12, lambda l: _SI_NO(l.has_replied)),
    ("Reunión agendada", 16, lambda l: _SI_NO(l.meeting_scheduled)),
    ("Creado", 18, lambda l: l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else ""),
]

_HEADER_FILL = PatternFill("solid", fgColor="0B2545")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_SI_FILL = PatternFill("solid", fgColor="D6F5E3")   # verde suave para "Sí"
_NO_FILL = PatternFill("solid", fgColor="F3F4F6")   # gris suave para "No"
_MEET_FILL = PatternFill("solid", fgColor="FFE9A8")  # dorado para reunión


def export_leads_to_excel(output_path: str = OUTPUT_PATH, ids: list[int] = None) -> str:
    """Construye el archivo Excel y devuelve la ruta donde se guardó."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Encabezados
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # Filas
    with session_scope() as db:
        query = db.query(Lead)
        if ids:
            query = query.filter(Lead.id.in_(ids))
        leads = query.order_by(Lead.created_at.desc()).all()
        row = 2
        for lead in leads:
            for col_idx, (_, _, extractor) in enumerate(COLUMNS, start=1):
                value = extractor(lead)
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")
                header = COLUMNS[col_idx - 1][0]
                # Coloreado de los estados Sí/No para lectura rápida.
                if header in ("Correo enviado", "Contestó"):
                    cell.fill = _SI_FILL if value == "Sí" else _NO_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header == "Reunión agendada":
                    cell.fill = _MEET_FILL if value == "Sí" else _NO_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        total = len(leads)

    # Resumen al pie
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"{total} leads").font = Font(bold=True)

    wb.save(output_path)
    logger.info("Reporte de Excel generado: %s (%s leads)", output_path, total)
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    path = export_leads_to_excel()
    print(f"Reporte generado en: {path}")
